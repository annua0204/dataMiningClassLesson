#!/usr/bin/python
# -*- coding: utf-8 -*-

import pymysql
import re
# from openpyxl import load_workbook
import openpyxl
from datetime import datetime

connection = pymysql.connect(host='127.0.0.1', port=3306, user='annua', passwd='Sorrow0081', db='data_mining', charset='utf8mb4')
cursor = connection.cursor(cursor=pymysql.cursors.DictCursor)

def change_datetime_string(cdata_string):
    datetime_string = None
    cdata_string = cdata_string.replace(" ", "")
    if (len(str(cdata_string)) == 7):
        cdata_string = str(cdata_string)
        year = str(int(cdata_string[0:3]) + 1911)
        month = cdata_string[3:5]
        date = cdata_string[5:7]
        dt = datetime.strptime(year+"-"+month+"-"+date, "%Y-%m-%d")
        datetime_string = dt.strftime("%Y-%m-%d")
    return datetime_string

def check_string(col, data_type):
    value = col.replace(" ", "");
    return value if (value != "") else None
    # if (data_type == "s"):
    #     return str(col) if (type(col) != unicode) else None
    # else:
    #     return col if (type(col) != unicode) else None

def getExcelDatas():
    print('into getExcelDatas()')
    print('start read data')
    wb = openpyxl.load_workbook("./2019_surger_datas_1001.xlsx")
    # wb = openpyxl.load_workbook("./2019_surger_datas_1007_short.xlsx")
    print('success read data')
    # 獲得所有sheet的名稱
    print(wb.get_sheet_names())
    # # 根據sheet名字獲得sheet
    sheet = wb.get_sheet_by_name('Row Data')
    import_data(sheet)

def import_data(sheet):
    print('into import_data()====')
    index = 0
    # 因為按行，所以返回A1, B1, C1這樣的順序
    for record in sheet.rows:
        if (record[2].value != 'ODR_TXDT'):
            data = {}
            index += 1
            print(index)
            data['ID'] = index
            data['ODR_LOGN'] =  check_string(record[0].value, 's')
            data['ODR_CHRT'] =  check_string(record[1].value, 's')
            data['ODR_TXDT'] = change_datetime_string(record[2].value)
            data['ODR_OPRM'] =  check_string(record[3].value, 's')
            data['ODR_DEPT'] =  check_string(record[4].value, 's')
            data['ODR_PSRC'] =  check_string(record[5].value, 's')
            data['ODR_BDNO'] =  check_string(record[6].value, 's')
            data['ODR_EFLG'] =  check_string(record[7].value, 's')
            data['ODR_IPNO'] =  check_string(record[8].value, 's')
            data['ODR_AS_D'] = change_datetime_string(record[9].value)
            data['ODR_AS_T'] =  check_string(record[10].value, 's')
            data['ODR_AE_D'] = change_datetime_string(record[11].value)
            data['ODR_AE_T'] =  check_string(record[12].value, 's')
            data['ODR_IN_D'] = change_datetime_string(record[13].value)
            data['ODR_IN_T'] =  check_string(record[14].value, 's')
            data['ODR_OS_D'] = change_datetime_string(record[15].value)
            data['ODR_OS_T'] =  check_string(record[16].value, 's')
            data['ODR_OE_D'] = change_datetime_string(record[17].value)
            data['ODR_OE_T'] =  check_string(record[18].value, 's')
            data['ODR_OT_D']=change_datetime_string(record[19].value)
            data['ODR_OT_T'] =  check_string(record[20].value, 's')
            data['ODR_OP_1'] =  check_string(record[21].value, 's')
            data['ODR_OP_2'] =  check_string(record[22].value, 's')
            data['ODR_OP_3'] =  check_string(record[25].value, 's')
            data['ODR_OP_4'] =  check_string(record[28].value, 's')
            data['ODR_M_DR'] =  check_string(record[31].value, 's')
            data['ODR_DN_1'] =  check_string(record[32].value, 's')
            data['ODR_DN_2'] =  check_string(record[33].value, 's')
            data['ODR_WN_1'] =  check_string(record[34].value, 's')
            data['ODR_WN_2'] =  check_string(record[35].value, 's')
            data['ODR_PAYK'] =  check_string(record[40].value, 's')
            data['ODR_AN_D'] =  check_string(record[43].value, 's')
            data['ODR_ASA'] = check_string(record[46].value, 'd')
            data['ODR_M_D2'] =  check_string(record[61].value, 's')
            data['ODR_M_D3'] =  check_string(record[62].value, 's')
            data['ODR_M_D4'] =  check_string(record[63].value, 's')
            data['ODR_WOUD'] = check_string(record[79].value, 'd')
            # print(data)
            sql_string = "INSERT INTO `2019_surger_datas` (`ID`, `ODR_LOGN`, `ODR_CHRT`, `ODR_TXDT`, `ODR_OPRM`, `ODR_DEPT`, `ODR_PSRC`, `ODR_BDNO`, `ODR_EFLG`, `ODR_IPNO`, `ODR_AS_D`, `ODR_AS_T`, `ODR_AE_D`, `ODR_AE_T`, `ODR_IN_D`, `ODR_IN_T`, `ODR_OS_D`, `ODR_OS_T`, `ODR_OE_D`, `ODR_OE_T`, `ODR_OT_D`, `ODR_OT_T`, `ODR_OP_1`, `ODR_OP_2`, `ODR_OP_3`, `ODR_OP_4`, `ODR_M_DR`, `ODR_DN_1`, `ODR_DN_2`, `ODR_WN_1`, `ODR_WN_2`, `ODR_PAYK`, `ODR_AN_D`, `ODR_ASA`, `ODR_M_D2`, `ODR_M_D3`, `ODR_M_D4`, `ODR_WOUD`) VALUES ( %(ID)s, %(ODR_LOGN)s ,%(ODR_CHRT)s ,%(ODR_TXDT)s ,%(ODR_OPRM)s ,%(ODR_DEPT)s ,%(ODR_PSRC)s ,%(ODR_BDNO)s ,%(ODR_EFLG)s ,%(ODR_IPNO)s ,%(ODR_AS_D)s ,%(ODR_AS_T)s ,%(ODR_AE_D)s ,%(ODR_AE_T)s ,%(ODR_IN_D)s ,%(ODR_IN_T)s ,%(ODR_OS_D)s ,%(ODR_OS_T)s ,%(ODR_OE_D)s ,%(ODR_OE_T)s ,%(ODR_OT_D)s ,%(ODR_OT_T)s ,%(ODR_OP_1)s ,%(ODR_OP_2)s ,%(ODR_OP_3)s ,%(ODR_OP_4)s ,%(ODR_M_DR)s ,%(ODR_DN_1)s ,%(ODR_DN_2)s ,%(ODR_WN_1)s ,%(ODR_WN_2)s ,%(ODR_PAYK)s ,%(ODR_AN_D)s ,%(ODR_ASA)s ,%(ODR_M_D2)s ,%(ODR_M_D3)s ,%(ODR_M_D4)s ,%(ODR_WOUD)s)"
            cursor.execute(sql_string, data)
            connection.commit()

    #关闭游标和数据库的连接
    cursor.close()
    connection.close()
    print('end')


getExcelDatas();
